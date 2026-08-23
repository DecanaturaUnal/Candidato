"""
Comprobaciones de las utilidades de operación.

No necesita credenciales ni conexión: sustituye la descarga de datos por un
conjunto de ejemplo y verifica que la exportación y el reporte funcionan, que las
tildes sobreviven y que no se puede colar una fórmula en la hoja de cálculo.

Uso (con el entorno virtual activado):
    python scripts/python/probar_utilidades.py
"""

from __future__ import annotations

import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import _comun  # noqa: E402

fallos = 0
total = 0


def comprobar(descripcion: str, condicion: bool, detalle: str = "") -> None:
    global fallos, total
    total += 1
    if condicion:
        print(f"  OK    {descripcion}")
    else:
        fallos += 1
        print(f"  FALLA {descripcion}" + (f"  -> {detalle}" if detalle else ""))


def comentarios_de_ejemplo() -> list[dict]:
    ahora = datetime.now(timezone.utc)
    return [
        {
            "id": "11111111-1111-4111-8111-111111111111",
            "nombre": "María Fernanda Ríos",
            "es_anonimo": False,
            "email": "maria.rios@example.com",
            "mensaje": "¿Qué plan hay para renovar los laboratorios?",
            "autoriza_publicacion": True,
            "autoriza_datos": True,
            "estado": "aprobado",
            "destacado": True,
            "likes_count": 7,
            "respuesta_decano": "Es una de las líneas del programa.",
            "respuesta_fecha": (ahora - timedelta(hours=2)).isoformat(),
            "created_at": (ahora - timedelta(days=3)).isoformat(),
            "moderado_por": "equipo@example.com",
            "moderado_en": (ahora - timedelta(days=2)).isoformat(),
        },
        {
            "id": "22222222-2222-4222-8222-222222222222",
            "nombre": "Anónimo",
            "es_anonimo": True,
            "email": "anonimo@example.com",
            # Empieza por "=": si no se escapa, Excel lo trata como fórmula.
            "mensaje": "=HYPERLINK(\"http://malicioso.example\",\"pulse aquí\")",
            "autoriza_publicacion": True,
            "autoriza_datos": True,
            "estado": "pendiente",
            "destacado": False,
            "likes_count": 0,
            "respuesta_decano": None,
            "respuesta_fecha": None,
            "created_at": (ahora - timedelta(hours=5)).isoformat(),
            "moderado_por": None,
            "moderado_en": None,
        },
    ]


EJEMPLOS = comentarios_de_ejemplo()

# Se sustituye la descarga real por los datos de ejemplo.
_comun.descargar_comentarios = lambda url, llave: list(EJEMPLOS)
_comun.cargar_entorno = lambda: ("https://ejemplo.supabase.co", "llave-de-prueba")

print("Escape de fórmulas:")
comprobar(
    "un texto que empieza por = se neutraliza",
    _comun.texto_seguro_para_hoja("=1+1").startswith("'"),
    _comun.texto_seguro_para_hoja("=1+1"),
)
for prefijo in ("+", "-", "@"):
    comprobar(
        f"lo mismo con «{prefijo}»",
        _comun.texto_seguro_para_hoja(f"{prefijo}algo").startswith("'"),
    )
comprobar(
    "un texto normal no se toca",
    _comun.texto_seguro_para_hoja("Hola, ¿qué tal?") == "Hola, ¿qué tal?",
)

print("\nExportación a Excel:")
import exportar_excel  # noqa: E402

exportar_excel.cargar_entorno = _comun.cargar_entorno
exportar_excel.descargar_comentarios = _comun.descargar_comentarios
exportar_excel.main()

from openpyxl import load_workbook  # noqa: E402

archivos = sorted(
    _comun.carpeta_de_salida().glob("comentarios_*.xlsx"),
    key=lambda p: p.stat().st_mtime,
)
comprobar("se generó el archivo", bool(archivos))

if archivos:
    hoja = load_workbook(archivos[-1]).active
    comprobar("tiene encabezado y 2 filas", hoja.max_row == 3, f"filas={hoja.max_row}")

    valores = [
        [celda.value for celda in fila] for fila in hoja.iter_rows(min_row=2)
    ]
    textos = [str(v) for fila in valores for v in fila if v is not None]

    comprobar(
        "las tildes sobreviven",
        any("María Fernanda Ríos" in t for t in textos),
    )
    comprobar(
        "la fórmula quedó neutralizada",
        any(t.startswith("'=HYPERLINK") for t in textos),
        next((t[:30] for t in textos if "HYPERLINK" in t), "no encontrada"),
    )
    comprobar(
        "los booleanos se leen en español",
        any(t == "Sí" for t in textos) and any(t == "No" for t in textos),
    )
    archivos[-1].unlink()

print("\nReporte:")
import reporte  # noqa: E402

reporte.cargar_entorno = _comun.cargar_entorno
reporte.descargar_comentarios = _comun.descargar_comentarios
try:
    reporte.main()
    comprobar("el reporte se genera sin errores", True)
except Exception as error:  # noqa: BLE001
    comprobar("el reporte se genera sin errores", False, str(error))

print(
    f"\n{total - fallos}/{total} comprobaciones correctas"
    + (f"  --  {fallos} FALLO(S)" if fallos else "")
)
sys.exit(1 if fallos else 0)
