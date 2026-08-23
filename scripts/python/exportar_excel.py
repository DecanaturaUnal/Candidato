"""
Exporta todos los comentarios a un archivo de Excel.

Es la versión completa del botón «Exportar CSV» del panel: incluye las mismas
columnas pero con formato, anchos de columna y una fila de encabezado fija, para
poder trabajar el archivo directamente.

Uso (con el entorno virtual activado):
    python scripts/python/exportar_excel.py
"""

from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from _comun import (
    cargar_entorno,
    carpeta_de_salida,
    descargar_comentarios,
    preparar_consola,
    texto_seguro_para_hoja,
)

# (título de la columna, campo en la base, ancho)
COLUMNAS = [
    ("Fecha", "created_at", 20),
    ("Nombre", "nombre", 26),
    ("Anónimo", "es_anonimo", 10),
    ("Correo", "email", 30),
    ("Mensaje", "mensaje", 70),
    ("Estado", "estado", 13),
    ("Destacado", "destacado", 11),
    ("Me gusta", "likes_count", 10),
    ("Autoriza publicar", "autoriza_publicacion", 16),
    ("Autoriza datos", "autoriza_datos", 14),
    ("Respuesta del decano", "respuesta_decano", 60),
    ("Fecha de respuesta", "respuesta_fecha", 20),
    ("Moderado por", "moderado_por", 28),
    ("Moderado en", "moderado_en", 20),
    ("ID", "id", 38),
]

AZUL = "243468"
TURQUESA = "3BA7B3"


def formatear(campo: str, valor) -> str | int:
    if campo in ("es_anonimo", "destacado", "autoriza_publicacion", "autoriza_datos"):
        return "Sí" if valor else "No"
    if campo == "likes_count":
        return int(valor or 0)
    if campo in ("created_at", "respuesta_fecha", "moderado_en") and valor:
        # De ISO 8601 a algo legible en la hoja
        try:
            momento = datetime.fromisoformat(str(valor).replace("Z", "+00:00"))
            return momento.astimezone().strftime("%Y-%m-%d %H:%M")
        except ValueError:
            return texto_seguro_para_hoja(valor)
    return texto_seguro_para_hoja(valor)


def main() -> None:
    preparar_consola()
    url, llave = cargar_entorno()

    print("Descargando comentarios…")
    comentarios = descargar_comentarios(url, llave)
    print(f"  {len(comentarios)} comentarios.")

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Comentarios"

    # Encabezado
    for indice, (titulo, _, ancho) in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=1, column=indice, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=AZUL)
        celda.alignment = Alignment(vertical="center")
        hoja.column_dimensions[get_column_letter(indice)].width = ancho

    hoja.freeze_panes = "A2"

    for fila, comentario in enumerate(comentarios, start=2):
        for columna, (_, campo, _) in enumerate(COLUMNAS, start=1):
            celda = hoja.cell(
                row=fila, column=columna, value=formatear(campo, comentario.get(campo))
            )
            if campo in ("mensaje", "respuesta_decano"):
                celda.alignment = Alignment(wrap_text=True, vertical="top")

        # Los pendientes se resaltan: son los que piden trabajo.
        if comentario.get("estado") == "pendiente":
            hoja.cell(row=fila, column=6).fill = PatternFill("solid", fgColor="FFF4D6")
        if comentario.get("destacado"):
            hoja.cell(row=fila, column=7).fill = PatternFill("solid", fgColor=TURQUESA)

    hoja.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLUMNAS))}{max(len(comentarios) + 1, 1)}"
    )

    marca = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d_%H%M")
    destino = carpeta_de_salida() / f"comentarios_{marca}.xlsx"
    libro.save(destino)

    print(f"Listo: {destino}")


if __name__ == "__main__":
    main()
